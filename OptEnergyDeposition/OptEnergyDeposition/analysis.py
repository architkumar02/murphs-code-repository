# coding: utf-8
from ROOT import TFile, TH1F, TCanvas
import os
from openpyxl import Workbook
import re
import matplotlib.pyplot as plt
import numpy as np

def GetRootFiles(path=os.getcwd(),gammaKey='Co60',neutronKey='neutron'):
  """ GetRootFiles

    Gets a list of root files, splitting up into gamma and neutrons
    according to gammaKey and neutronKey
  """
  gammaFiles = list()
  neutronFiles = list()
  for filename in os.listdir(path):
    [filename,ext] = os.path.splitext(filename)
    if ext == '.root':
      particleName = filename.split('_')[0].split('run')[0]
      filename = os.path.join(path,filename)+ext
      if particleName == gammaKey:
        gammaFiles.append(filename)
      elif particleName == neutronKey:
        neutronFiles.append(filename)
      else:
        print 'ROOT file '+filename+' is not reconized'
  return [gammaFiles,neutronFiles]

def PrintFiles(filelist,wb=Workbook(),histKey='eDepHist'):
  """ PlotFiles
  Prints the files files contained in filelist
  Keywords:
  fileName - name of the figure to save
  histKey - Key of the histogram in the root file
  """
  longName = ['Energy','Frequency','Error']
  units =['MeV','Particles per Event','Particles per Event']
  row = 0
  col = 0
  with open(fileName,'w') as outfile:
    for fname in filelist:
      f = TFile(fname,'r')
      hist = f.Get(histKey)
      # Writing the header

def GetThickness(filename):
  """ 
  GetThickness
  
  Parses a filename for a given thickness, returns the thickness in mm
  """
  value = filename.split('_')[-1]
  value = value.split('.')[0]
  # Uses a regular expression to seperate
  pattern = re.compile('(\d+)(\w+)')
  tokens = pattern.split(value)
  if tokens[2] == 'm':
    return float(tokens[1])*1000
  elif tokens[2] == 'cm':
    return float(tokens[1])*10
  elif tokens[2] == 'mm':
    return float(tokens[1])
  elif tokens[2] == 'um':
    return float(tokens[1])*0.001
  else:
    print tokens[2]," is not a reconized prefix"
    raise Exception()

def PrintEDepSummary(gFiles,nFiles,wb=Workbook(),tParse=GetThickness,
  histKey='eDepHist'):
  """ PrintEDepSummary
  Prints the energy deposition summary
  """
  ws = wb.create_sheet()
  ws.title = 'EDepSummary'
  # Extrating the average values
  row = 0
  col = 0
  # Writing the headers
  longname = ['thickness','Average Energy Deposition','Avg. EDep Error']
  units =['mm','MeV','MeV']
  for l,u in zip(longname,units):
    ws.cell(row=0,column=col).value = l
    ws.cell(row=1,column=col).value = u
    ws.cell(row=0,column=col+3).value = l
    ws.cell(row=1,column=col+3).value = u
    col += 1
  ws.cell(row=2,column=1).value = 'Co-60'
  ws.cell(row=2,column=4).value = 'Cf-252'
  # Writing the data
  row = 3
  for fname in gFiles:
    f = TFile(fname,'r')
    hist = f.Get(histKey)
    ws.cell(row=row,column=0).value = GetThickness(fname)
    ws.cell(row=row,column=1).value = hist.GetMean()
    ws.cell(row=row,column=2).value = hist.GetMeanError()
    row += 1
  row = 3
  for fname in nFiles:
    f = TFile(fname,'r')
    hist = f.Get(histKey)
    ws.cell(row=row,column=3).value = GetThickness(fname)
    ws.cell(row=row,column=4).value = hist.GetMean()
    ws.cell(row=row,column=5).value = hist.GetMeanError()
    row += 1

def PlotEDepSummary(gFiles,nFiles,figureName='EDepSummary.png',tParse=GetThickness,
  histKey='eDepHist'):
  """ PlotEDepSummary
  Plotss the energy deposition summary
  """
  # Extrating the average values
  gT = list()
  gDep = list()
  gDepError = list()
  nT = list()
  nDep = list()
  nDepError = list()
  for fname in gFiles:
    f = TFile(fname,'r')
    hist = f.Get(histKey)
    gT.append(GetThickness(fname))
    gDep.append(hist.GetMean())
    gDepError.append(hist.GetMeanError())
  for fname in nFiles:
    f = TFile(fname,'r')
    hist = f.Get(histKey)
    nT.append(GetThickness(fname))
    nDep.append(hist.GetMean())
    nDepError.append(hist.GetMeanError())
  # Plotting
  plt.errorbar(gT,gDep,yerr=gDepError,fmt='ro')
  plt.hold(True)
  plt.errorbar(nT,nDep,yerr=nDepError,fmt='go')
  plt.xlabel("Thickness (mm)")
  plt.ylabel("Average Energy Deposition (MeV)")
  #plt.legend(["Co-60","Cf-252"])
  plt.xscale("log")
  plt.yscale("log")
  plt.grid(True)
  plt.savefig(figureName)

def main():
  print "Getting Files"
  [g,n] = GetRootFiles()
  print "Starting Data Analysis"
  wb = Workbook()
  PrintEDepSummary(g,n,wb)
  PlotEDepSummary(g,n)
  #print "Print Data"
  #PrintFiles(g,'GammaData.dat')
  wb.save('EnergyDepAnalysis.xlsx')

if __name__ == "__main__":
  main()
